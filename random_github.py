import random
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta

# List of site names
locations = ['London', 'Athens', 'Paris', 'Porto', 'Madrid', 'Prague', 'AFTERNOON']

# List of names from two groups
DRIVERS = ['DIM T', 'CHRIS P', 'MARIOS', 'AKIS', 'ANTONY', 
           'GEORGE', 'CHRIS M', 'DIM S', 'NICK', 'PANOS']
HEALTH_CARES = ['MARY K', 'MARY M', 'CHRISTINE', 'JOHANA', 'KELLY', 'ANGELA', 'TASOS', 'GABRIEL', 'SYLVIE',
'DIM M', 'MARY G', 'ATHANASIA', 'KATHRINE M', 'VASO', 'FOTIS', 'HELEN', 'MARY F', 'CHRIS MP', 'LYDIA', 'MARY KO', 'MATINA']

# Add the new names that will be displayed permanently
ADDITIONAL_NAMES = {
    'MARY G': 'SPECIAL CARE',
    'KATHRINE K': 'REPOS',
    'RITA': 'ON VACATIONS'
}

# Days of the week
days = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY']

# Calculating next week's dates
today = datetime.today()
next_monday = today + timedelta(days=(7 - today.weekday() + 0) % 7)
dates = [(next_monday + timedelta(days=i)).strftime('%d/%m/%Y') for i in range(5)]

# Create a dictionary for the project
schedule = {day: {loc: [] for loc in locations} for day in days}

# Create a dictionary to keep track of where each name has gone
name_to_locations = {name: [] for name in DRIVERS + HEALTH_CARES}
name_to_days = {name: {day: None for day in days} for name in DRIVERS + HEALTH_CARES}

# Receive user input for the evening shift
print("gIVE THE NAMES FOR AFTERNOON WORK.")
afternoon_name1 = input("Give the name from list 1 for the afternoon: ")
afternoon_names2 = [input(f"Give the name from list 2 for the afternoon: (NAME {i+1}): ") for i in range(3)]

# Adding the evening shift for each day
for day in days:
    schedule[day]['AFTERNOON'] = [afternoon_name1] + afternoon_names2
    name_to_locations[afternoon_name1].append('AFTERNOON')
    name_to_days[afternoon_name1][day] = 'AFTERNOON'
    for name in afternoon_names2:
        name_to_locations[name].append('AFTERNOON')
        name_to_days[name][day] = 'AFTERNOON'

# Function to find a valid random combination of names for a location
def get_valid_names(loc, current_names, num_needed, name_list, day):
    chosen_names = []
    attempts = 0
    while len(chosen_names) < num_needed and attempts < 100:
        name = random.choice(name_list)
        if (name not in current_names and 
            loc not in name_to_locations[name] and 
            name not in chosen_names and 
            name_to_days[name][day] is None):
            chosen_names.append(name)
        attempts += 1
    return chosen_names

# Distribution of sites and names for the remaining sites
for day in days:
    random.shuffle(locations)
    
    for loc in locations:
        if loc == 'AFTERNOON':
            continue
        
        # We try to find the required names according to the list to be created
        if loc in ['London', 'Athens', 'Paris', 'Porto', 'Madrid']:
            num_names1 = 2
            num_names2 = 3
        if loc in ['Prague']:
            num_names1 = 1
            num_names2 = 4

        # Finding the required names
        chosen_names1 = get_valid_names(loc, schedule[day][loc], num_names1, DRIVERS, day)
        chosen_names2 = get_valid_names(loc, schedule[day][loc] + chosen_names1, num_names2, HEALTH_CARES, day)
        
        # Adding names to the program and updating the dictionaries
        schedule[day][loc].extend(chosen_names1)
        schedule[day][loc].extend(chosen_names2)
        
        for name in chosen_names1 + chosen_names2:
            name_to_locations[name].append(loc)
            name_to_days[name][day] = loc

# Ensuring that all sites are covered for each day
for day in days:
    for loc in locations:
        if loc != 'AFTERNOON' and not schedule[day][loc]:
            chosen_names1 = get_valid_names(loc, schedule[day][loc], 2, DRIVERS, day)
            chosen_names2 = get_valid_names(loc, schedule[day][loc] + chosen_names1, 3, HEALTH_CARES, day)
            schedule[day][loc].extend(chosen_names1)
            schedule[day][loc].extend(chosen_names2)
            for name in chosen_names1 + chosen_names2:
                name_to_locations[name].append(loc)
                name_to_days[name][day] = loc

# Create DataFrame for export to Excel
df = pd.DataFrame(index=DRIVERS + HEALTH_CARES + list(ADDITIONAL_NAMES.keys()), columns=days)

# Add the programmes to the corresponding columns
for day in days:
    for loc in locations:
        for name in schedule[day][loc]:
            df.loc[name, day] = loc

# Addition of permanent names and phrases
for name, phrase in ADDITIONAL_NAMES.items():
    df.loc[name] = [phrase] * len(days)

# Export the DataFrame to Excel
excel_filename = 'random_programme.xlsx'
df.to_excel(excel_filename, index=True)

# Add colors and fonts to your Excel file
wb = load_workbook(excel_filename)
ws = wb.active

# Add dates on the second line
ws.insert_rows(1)
for col_num, date in enumerate(dates, 2):
    ws.cell(row=1, column=col_num, value=date)

# Dictionary of colors for each location
color_fills = {
    "London": PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid"),
    "Athens": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "Paris": PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid"),
    "Porto": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "Madrid": PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid"),
    "Prague": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
}

# Application of colours and font
for row in ws.iter_rows(min_row=3, min_col=2):  # Adaptation to the new lines
    for cell in row:
        location = cell.value
        if location in color_fills:
            cell.fill = color_fills[location]
        if location == 'AFTERNOON':
            cell.font = Font(color="0000FF")

# Adjust the width of the columns
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the Excel file with the adjustments
wb.save(excel_filename)

print(f"The program was saved in the {excel_filename} file with the dates, colours and fonts you requested.")
